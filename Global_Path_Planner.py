from CV_Module import Global_Planner_CV_Module
from Trans_Module import Global_Planner_Trans_Module
# from NeurophysicalModel_Module import Global_Planner_NeurophysicalModel_Module
import math
import numpy as np
import cv2
import pickle
import openpyxl
import sys


class Global_Path_Planner():
    def __init__(self, imgname):
        self.CV_Module = Global_Planner_CV_Module(imgname)
        self.Trans_Module = Global_Planner_Trans_Module()
        # self.NeurophysicalModel_Module = Global_Planner_NeurophysicalModel_Module()

    def __one_point_correction_1iter(self, speed, surfacetypes, disctime, start_point, end_point, epsilon=0.01):
        speed_0_1 = speed
        s_x, s_y = start_point[0], start_point[1]
        e_x, e_y = end_point[0], end_point[1]
        omniwheels_speed = self.Trans_Module.calculate_omni_wheel_velocities(10 * speed_0_1 * math.pi / 30)
        poses_dx, poses_dy, _, current, _ = self.NeurophysicalModel_Module.predict(omniwheels_speed[0],
                                                                                   omniwheels_speed[1],
                                                                                   omniwheels_speed[2],
                                                                                   surfacetypes[0],
                                                                                   surfacetypes[1],
                                                                                   surfacetypes[2],
                                                                                   disctime,
                                                                                   0)
        r_x_0_1, r_y_0_1 = s_x + poses_dx, s_y + poses_dy
        total_current_0_1 = np.sum(np.array([math.fabs(x) for x in current[0]]))
        # error = math.sqrt((e_x - r_x) ** 2 + (e_y - r_y) ** 2)
        print(s_x, s_y, "START")
        print(r_x_0_1, r_y_0_1, "REAL")
        print(e_x, e_y, "END")
        speed_0_2 = [(e_x - r_x_0_1) / 0.5, (e_y - r_y_0_1) / 0.5, 0]
        omniwheels_speed = self.Trans_Module.calculate_omni_wheel_velocities(10 * speed_0_2 * math.pi / 30)
        poses_dx, poses_dy, _, current, _ = self.NeurophysicalModel_Module.predict(omniwheels_speed[0],
                                                                                   omniwheels_speed[1],
                                                                                   omniwheels_speed[2],
                                                                                   surfacetypes[0],
                                                                                   surfacetypes[1],
                                                                                   surfacetypes[2],
                                                                                   disctime,
                                                                                   0)
        r_x_0_2, r_y_0_2 = r_x_0_1 + poses_dx, r_y_0_1 + poses_dy
        total_current_0_2 = np.sum(np.array([math.fabs(x) for x in current[0]]))
        print(r_x_0_1, r_y_0_1, "START")
        print(r_x_0_2, r_y_0_2, "REAL")
        print(e_x, e_y, "END")

        return [r_x_0_2, r_y_0_2], total_current_0_1 + total_current_0_2

    def __one_point_correction_while(self, speed, surfacetypes, disctime, start_point, end_point, epsilon=0.01):
        speed = self.Trans_Module.vel_ps_2_ms(speed)
        s_x, s_y = self.Trans_Module.coords_from_p_2_m(start_point[0], start_point[1])
        e_x, e_y = self.Trans_Module.coords_from_p_2_m(end_point[0], end_point[1])
        additional_speed = [0, 0, 0]
        error = 1
        while error > epsilon:
            speed_corrected = [speed[i] + additional_speed[i] for i in range(len(speed))]
            omniwheels_speed = self.Trans_Module.calculate_omni_wheel_velocities(speed_corrected)
            poses_dx, poses_dy, _, _, _ = self.NeurophysicalModel_Module.predict(omniwheels_speed[0],
                                                                                 omniwheels_speed[1],
                                                                                 omniwheels_speed[2],
                                                                                 surfacetypes[0],
                                                                                 surfacetypes[1],
                                                                                 surfacetypes[2],
                                                                                 disctime,
                                                                                 0)
            r_x, r_y = s_x + poses_dx, s_y + poses_dy
            error = math.sqrt((e_x - r_x) ** 2 + (e_y - r_y) ** 2)
            additional_speed = [e_x - r_x, e_y - r_y, 0]
            print(s_x, s_y, "Start")
            print(r_x, r_y, "Real")
            print(e_x, e_y, "End")
            print(error, "error")
            print(speed_corrected, "speed ms")
            print(additional_speed, "additional_speed")

    def __generate_speeds(self, speed_magnitude, angle_step):  # Генерирует скорости для одной ампилутды
        speeds = []

        for angle in range(0, 360, angle_step):
            angle_rad = math.radians(angle)

            vx = speed_magnitude * math.cos(angle_rad)
            vy = speed_magnitude * math.sin(angle_rad)
            wz = 0.0

            speed = np.array([vx, vy, wz])
            speeds.append(speed)

        return speeds

    def __generate_speeds_full(self, angle_step):  # Генерирует скорости для трех амплитуд 0.1 0.2 0.3
        speeds = []
        amps = [0.1, 0.2, 0.3]

        for amp in amps:
            for angle in range(0, 360, angle_step):
                angle_rad = math.radians(angle)

                vx = amp * math.cos(angle_rad)
                vy = amp * math.sin(angle_rad)
                wz = 0.0

                speed = np.array([vx, vy, wz])
                speeds.append(speed)

        return speeds

    def __generate_types(self):
        values = [2, 3]
        types = []

        for x1 in values:
            for x2 in values:
                for x3 in values:
                    types.append((x1, x2, x3))

        return types

    def __generate_extra_types(self):
        values = [2, 3, 4]
        types = []

        for x1 in values:
            for x2 in values:
                for x3 in values:
                    if 4 in (x1, x2, x3):
                        types.append((x1, x2, x3))

        return types

    def generate_combs(self, angle_disct):
        filename = "combs_dict_v_2_0.pkl"
        speeds = self.__generate_speeds_full(angle_disct)
        types = self.__generate_types()
        extra_types = self.__generate_extra_types()
        print(extra_types)
        dict_tmp = {(0, 0, 0): None}
        i = 0
        print(len(types))
        print(len(extra_types))
        print(len(speeds))
        for type_tmp in types:
            tmp = []

            for speed in speeds:
                omniwheels_speeds = self.Trans_Module.calculate_omni_wheel_velocities(speed)
                data = self.NeurophysicalModel_Module.predict(round(omniwheels_speeds[0] * 10 * math.pi / 30, 2),
                                                              round(omniwheels_speeds[1] * 10 * math.pi / 30, 2),
                                                              round(omniwheels_speeds[2] * 10 * math.pi / 30, 2),
                                                              type_tmp[0],
                                                              type_tmp[1],
                                                              type_tmp[2], 0.5, 0)
                data.append(omniwheels_speeds)
                tmp.append(data)
                print(data)
                i += 1
                print(i)

            dict_tmp[type_tmp] = tmp

        i = 0

        with open(filename, 'wb') as file:
            pickle.dump(dict_tmp, file)
        return

    def __find_optimal_speed(self, start_point, end_point, available_speeds, disctime):
        min_distance = float('inf')
        optimal_speed = None

        for speed in available_speeds:
            delta = [speed[0] * disctime, speed[1] * disctime]
            new_point = np.add(start_point, delta)

            distance = np.linalg.norm(np.subtract(new_point, end_point))

            if distance < min_distance:
                min_distance = distance
                optimal_speed = speed

        return optimal_speed

    def __distance(self, start, end):
        return math.sqrt(((start[0] - end[0]) ** 2 + (start[1] - end[1]) ** 2))

    def load_combs(self, filename):
        with open(filename, 'rb') as file:
            return pickle.load(file)

    def calc_score1(self, dist, dist_new, pred):
        difference_dist = lambda x, y: 1 if x - y > 0 else -1
        return 1 / (difference_dist(dist, dist_new) * sum(map(abs, pred[3][0])))

    def calc_score2(self, dist, dist_new, pred):
        return (dist - dist_new) / (sum(map(abs, pred[3][0])))

    def calc_score3(self, dist, dist_new, pred):
        difference_dist = lambda x, y: 1 if x - y > 0 else -1
        tmp = [pred[3][0][i] * pred[-1][i] for i in range(3)]
        total_watt = sum(map(abs, tmp))
        return 1 / (difference_dist(dist, dist_new) * total_watt)

    def calc_score4(self, dist, dist_new, pred):
        tmp = [pred[3][0][i] * pred[-1][i] for i in range(3)]
        total_watt = sum(map(abs, tmp))
        return (dist - dist_new) / total_watt

    def get_new_pos_n_score(self, filename, type_wheels, test_start, test_end, angle):
        combs_dict = self.load_combs(filename)
        scores = []
        for pred in combs_dict[type_wheels]:
            dist = math.sqrt((test_end[0] - test_start[0]) ** 2 + (test_end[1] - test_start[1]) ** 2)
            dx, dy = (pred[0] * math.cos(angle) + pred[1] * math.sin(angle)), (
                    pred[1] * math.cos(angle) + pred[0] * math.sin(angle))
            test_point = [test_start[0] + dy, test_start[1] + dx]
            dist_new = math.sqrt((test_end[0] - test_point[0]) ** 2 + (test_end[1] - test_point[1]) ** 2)
            scores.append(self.calc_score4(dist, dist_new, pred))
        max_positive_value, max_positive_index = max((x, i) for i, x in enumerate(scores) if x > 0)
        dx, dy = (combs_dict[type_wheels][max_positive_index][0] * math.cos(angle) +
                  combs_dict[type_wheels][max_positive_index][1] * math.sin(angle)), (
                combs_dict[type_wheels][max_positive_index][1] * math.cos(angle) +
                combs_dict[type_wheels][max_positive_index][0] * math.sin(angle))
        new_point = [test_start[0] + dy,
                     test_start[1] + dx]
        return max_positive_value, combs_dict[type_wheels][max_positive_index][3], new_point, combs_dict[type_wheels][
            max_positive_index]

    def save_to_excel(self, file_path, row, col1, col2, col3, value1, value2, value3):
        wb = openpyxl.load_workbook(file_path)
        # Выбор активного листа (или можно выбрать любой другой лист)
        ws = wb.active

        # Присваиваем значения ячейкам
        ws[f'{col1}{row}'] = value1
        ws[f'{col2}{row}'] = value2
        ws[f'{col3}{row}'] = value3

        # Сохраняем изменения в файл
        wb.save(file_path)
        wb.close()

    def save_to_pkl(self, data, filename):
        filename1 = filename + "_speeds.pkl"
        with open(filename1, 'wb') as file:
            pickle.dump(data, file)

    def calculate_traj(self, speed, disctime, start, end, number, epsilon=0.03):
        result_points, result_vels, surface_types_per_wheels, img, path, contmass, contscenters = self.CV_Module.points_and_vels(
            speed, disctime,
            start,
            end, number)
        print("Path: ", path)
        print("Contscenters: ", contscenters)
        start_point_m = self.Trans_Module.coords_from_p_2_m(contscenters[path[0]][0], contscenters[path[0]][1])
        start_point_p = contscenters[path[0]]
        print("____________________________________________________________________")
        print("_____________________________PLANNER________________________________")
        print(start_point_m, "START_POINT")
        end_point_m = self.Trans_Module.coords_from_p_2_m(contscenters[path[-1]][0], contscenters[path[-1]][1])
        end_point_p = contscenters[path[-1]]
        print(end_point_m, "END_POINT")
        # surface_types = self.CV_Module.find_wheels_pos(start_point_p, contmass)
        # dist = self.__distance(start_point_m, end_point_m)
        point_tmp = start_point_m
        angle = 0
        global_score = 0
        global_current = 0
        global_watt = 0
        global_efficient = 0
        points = []
        points_movement_data = []
        length_total = 0
        for j in range(1, len(path)):
            end_point_m = self.Trans_Module.coords_from_p_2_m(contscenters[path[j]][0], contscenters[path[j]][1])
            dist = self.__distance(start_point_m, end_point_m)
            total_score = 0

            point_tmp = start_point_m
            v = 0
            while dist > epsilon:
                v += 1
                points.append(point_tmp)
                surface_types = self.CV_Module.find_wheels_pos(self.Trans_Module.coords_from_m_2_p_mass(point_tmp),
                                                               contmass, angle)
                score, pred_info, point_tmp, motion_data = self.get_new_pos_n_score("combs_dict_amp_full.pkl",
                                                                                    tuple(surface_types),
                                                                                    point_tmp,
                                                                                    end_point_m, angle)
                angle += motion_data[2]
                points_movement_data.append(motion_data)
                global_watt += sum([abs(motion_data[-1][i] * pred_info[0][i]) for i in range(3)])
                global_current += sum(abs(pred_info[0]))
                dist_prev = dist
                dist = self.__distance(point_tmp, end_point_m)
                global_efficient += (sum([abs(motion_data[-1][i] * pred_info[0][i]) for i in range(3)])) / (
                        (dist_prev - dist) * 1000)
                print(self.Trans_Module.calculate_vxvywz_from_omniwheels_vels(motion_data[-1]))
                total_score += score
                print(f"_____________________________ITER {v}________________________________")
                print("NEW POINT: ", point_tmp)
                print("DIST", dist)
                print("SCORE", score)
                print("TOTAL CURRENT", global_current)
                print("TOTAL WATT", str(global_watt).replace('.', ','))
                print("DIST PREV", dist_prev)
                print("DIST", dist)
                print("TOTAL EFFICIENT", str(1 / global_efficient * 10000).replace('.', ','))
                print("MOTION DATA: ", motion_data)
                print("SPEEDS VX VY WZ: ", self.Trans_Module.calculate_vxvywz_from_omniwheels_vels(motion_data[-1]))
                print("ANGLE: ", angle)
                # print(motion_data)

            start_point_m = point_tmp
            global_score += total_score
        for i in range(1, len(points)):
            cv2.circle(img, self.Trans_Module.coords_from_m_2_p_mass(points[i]), 3, (0, 0, 255), 2)
            cv2.line(img, self.Trans_Module.coords_from_m_2_p_mass(points[i - 1]),
                     self.Trans_Module.coords_from_m_2_p_mass(points[i]), (0, 255, 255), 2)
        print("LENGTH OF PATH: ", str(self.calculate_dijkstra_path_length(path, contscenters)).replace('.', ','))
        print("LENGTH OF EDITED PATH: ", str(self.calculate_path_length_from_scored(points)).replace('.', ','))
        print("NUMBER OF PATH: ", number)
        # self.save_to_excel('Stockpile/AMPS_FULL_w_angle/output_edited_w_corrected_angle.xlsx', 3 + (12 * 1) + number, 'M', 'N', 'O',
        #                    round(global_watt, 2), round((1 / global_efficient * 10000), 2),
        #                    round(self.calculate_path_length_from_scored(points), 2))
        # print("_________________Speeds_________________")
        # speeds = []
        # for data in points_movement_data:
        #     speed = self.Trans_Module.calculate_vxvywz_from_omniwheels_vels(data[-1])
        #     speed[-1] = 0
        #     speeds.append(list(speed))
        # print(str(speeds).replace('[', '{').replace(']', '}'))
        # self.save_to_pkl(points_movement_data, f'Stockpile/AMPS_FULL/Movement_data_f_pr_test/{number}_length_S4')
        cv2.imwrite(f"Stockpile/AMPS_FULL_w_angle/length_only/{number}_length_only_S4.png", img)

    def calculate_dijkstra_path_length(self, path, contscenters):
        path_m = [self.Trans_Module.coords_from_p_2_m_mass(contscenters[i]) for i in path]
        length = 0
        for i in range(1, len(path_m)):
            length += math.sqrt((path_m[i][0] - path_m[i - 1][0]) ** 2 + (path_m[i][1] - path_m[i - 1][1]) ** 2)

        return length

    def calculate_path_length_from_scored(self, points):
        length = 0
        for i in range(1, len(points)):
            length += math.sqrt((points[i][0] - points[i - 1][0]) ** 2 + (points[i][1] - points[i - 1][1]) ** 2)
        return length

    def calculate_traj_w_edited(self, speed, disctime, start, end, number, epsilon=0.03):
        result_points, result_vels, _, img, path, contmass, contscenters = self.CV_Module.points_and_vels(
            speed, disctime,
            start,
            end, number)

        start_point_m = self.Trans_Module.coords_from_p_2_m(contscenters[path[0]][0], contscenters[path[0]][1])
        start_point_p = contscenters[path[0]]
        print("____________________________________________________________________")
        print("_____________________________PLANNER________________________________")
        print(start_point_m, "START_POINT")
        end_point_m = self.Trans_Module.coords_from_p_2_m(contscenters[path[-1]][0], contscenters[path[-1]][1])
        end_point_p = contscenters[path[-1]]
        print(end_point_m, "END_POINT")
        # surface_types = self.CV_Module.find_wheels_pos(start_point_p, contmass)
        # dist = self.__distance(start_point_m, end_point_m)
        point_tmp = start_point_m
        global_score = 0
        global_current = 0
        global_watt = 0
        global_efficient = 0
        points = []
        for j in range(1, len(path)):
            end_point_m = self.Trans_Module.coords_from_p_2_m(contscenters[path[j]][0], contscenters[path[j]][1])
            dist = self.__distance(start_point_m, end_point_m)
            total_score = 0

            point_tmp = start_point_m
            v = 0
            while dist > epsilon:
                v += 1
                points.append(point_tmp)
                surface_types = self.CV_Module.find_wheels_pos_edited(
                    self.Trans_Module.coords_from_m_2_p_mass(point_tmp),
                    contmass)
                print("SURFACE TYPES: ", surface_types)
                score, pred_info, point_tmp, motion_data = self.get_new_pos_n_score("combs_dict_amp_full_edited.pkl",
                                                                                    tuple(surface_types),
                                                                                    point_tmp,
                                                                                    end_point_m)
                global_watt += sum([abs(motion_data[-1][i] * pred_info[0][i]) for i in range(3)])
                global_current += sum(abs(pred_info[0]))
                dist_prev = dist
                dist = self.__distance(point_tmp, end_point_m)
                global_efficient += (sum([abs(motion_data[-1][i] * pred_info[0][i]) for i in range(3)])) / (
                        (dist_prev - dist) * 1000)

                total_score += score
                print(f"_____________________________ITER {v}________________________________")
                print("DIST", dist)
                print("SCORE", score)
                print("TOTAL CURRENT", global_current)
                print("TOTAL WATT", global_watt)
                print("DIST PREV", dist_prev)
                print("DIST", dist)
                print("TOTAL EFFICIENT", 1 / global_efficient * 10000)
                # print(motion_data)

            start_point_m = point_tmp
            global_score += total_score
        for i in range(1, len(points)):
            cv2.circle(img, self.Trans_Module.coords_from_m_2_p_mass(points[i]), 3, (0, 0, 255), 2)
            cv2.line(img, self.Trans_Module.coords_from_m_2_p_mass(points[i - 1]),
                     self.Trans_Module.coords_from_m_2_p_mass(points[i]), (0, 255, 255), 2)

        cv2.imwrite(f"Stockpile/AMPS_FULL/coeffs_only/{number}_coeff_only_S1.png", img)

    def calculate_traj_start_end(self, speed, disctime, start, end, number,
                                 epsilon=0.03):  # Реализует маршрут от нач точки до кон точкиы
        result_points, result_vels, surface_types_per_wheels, img, path, contmass, contscenters = self.CV_Module.points_and_vels(
            speed, disctime,
            start,
            end, number)
        img = cv2.imread("1.jpg_croped.png_done.png")
        start_point_m = self.Trans_Module.coords_from_p_2_m(contscenters[path[0]][0], contscenters[path[0]][1])
        start_point_p = contscenters[path[0]]
        print("____________________________________________________________________")
        print("_____________________________PLANNER________________________________")
        print(start_point_m, "START_POINT")
        end_point_m = self.Trans_Module.coords_from_p_2_m(contscenters[path[-1]][0], contscenters[path[-1]][1])
        end_point_p = contscenters[path[-1]]
        print(end_point_m, "END_POINT")
        # surface_types = self.CV_Module.find_wheels_pos(start_point_p, contmass)
        # dist = self.__distance(start_point_m, end_point_m)
        point_tmp = start_point_m
        global_current = 0
        dist = self.__distance(start_point_m, end_point_m)
        v = 0
        points = []
        while dist > epsilon:
            v += 1
            points.append(point_tmp)
            surface_types = self.CV_Module.find_wheels_pos_edited(self.Trans_Module.coords_from_m_2_p_mass(point_tmp),
                                                                  contmass)
            score, pred_info, point_tmp, _ = self.get_new_pos_n_score("combs_dict_amp_full_edited.pkl",
                                                                      tuple(surface_types),
                                                                      point_tmp,
                                                                      end_point_m)
            global_current += sum(abs(pred_info[0]))
            dist = self.__distance(point_tmp, end_point_m)
            print(f"_____________________________ITER {v}________________________________")
            print("DIST", dist)
            print("SCORE", score)
            print("TOTAL CURRENT", global_current)

        for i in range(1, len(points)):
            cv2.circle(img, self.Trans_Module.coords_from_m_2_p_mass(points[i]), 3, (0, 0, 255), 2)
            cv2.line(img, self.Trans_Module.coords_from_m_2_p_mass(points[i - 1]),
                     self.Trans_Module.coords_from_m_2_p_mass(points[i]), (0, 255, 255), 2)

        cv2.imwrite(f"{number}_score_sqrt_end_start_edited.png", img)

    def is_close(self, pos1, pos2, epsilon):
        # Проверка, находятся ли две точки достаточно близко друг к другу
        return abs(pos1[0] - pos2[0]) < epsilon and abs(pos1[1] - pos2[1]) < epsilon

    def get_surface_types(self, pos, contmass):
        # Определение типов поверхности под колесами в заданной позиции
        point_p = self.Trans_Module.coords_from_m_2_p_mass(pos)
        return self.CV_Module.find_wheels_pos(point_p, contmass)

    def move(self, pos, dx, dy):
        return [pos[0] + dx, pos[1] + dy]

    def calculate_total_current(self, route):
        # Вычисление суммарного тока для заданного маршрута
        total_current = 0
        for movement in route:
            currents = movement[3][0]  # Предполагается, что I1, I2, I3 - это токи на каждом из двигателей
            total_current += sum(abs(currents))
        return total_current

    def backtrack(self, current_route, current_pos, end_m, contmass, total_current, combs_dict, epsilon=0.03):
        global best_route, best_current
        if total_current > 170:
            return
        if self.is_close(current_pos, end_m, epsilon):
            total_current = self.calculate_total_current(current_route)
            if total_current < best_current:
                best_current = total_current
                best_route = current_route.copy()
            return
        current_surface_types = self.get_surface_types(current_pos, contmass)
        for movement in combs_dict[tuple(current_surface_types)]:
            dx, dy, phi, currents, _ = movement
            new_pos = self.move(current_pos, dx, dy)
            if new_pos is not None:
                current_route.append(movement)
                self.backtrack(current_route, new_pos, end_m, contmass, total_current + sum(abs(currents[0])),
                               combs_dict)
                current_route.pop()

    def bruteforce(self, dictfile_name, speed, disctime, start, end, number, epsilon=0.03):
        result_points, result_vels, surface_types_per_wheels, img, path, contmass, contscenters = self.CV_Module.points_and_vels(
            speed, disctime,
            start,
            end, number)

        combs_dict = self.load_combs(dictfile_name)
        start_p = contscenters[start]
        end_p = contscenters[end]
        start_m = self.Trans_Module.coords_from_p_2_m_mass(start_p)
        end_m = self.Trans_Module.coords_from_p_2_m_mass(end_p)

        best_route = None
        best_current = float('inf')
        print("START!!!")
        self.backtrack([], start_m, end_m, contmass, 0, combs_dict)

        return best_route, best_current


# if __name__ == "__main__":
#     print(sys.argv[1])
#     index = int(sys.argv[1])  # Получение аргумента командной строки
#     Path_Planner = Global_Path_Planner("C:/Users/LostRiboxX/Desktop/Py/DiplomaWork/1.jpg")
#     Path_Planner.calculate_traj(51.5, 0.5, 13, 1, index)
Path_Planner = Global_Path_Planner("C:/Users/LostRiboxX/Desktop/Py/DiplomaWork/1.jpg")
Path_Planner.calculate_traj(51.5, 0.5, 13, 1, 3)
